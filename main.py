import os
import sys
from pathlib import Path
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
from docxtpl import DocxTemplate
from openpyxl import load_workbook


class Main(QDialog):
    def __init__(self):
        super(Main, self).__init__()
        loadUi('main.ui', self)
        self.setWindowTitle('Работа с файлами в Python')

        self.btn_wrd.clicked.connect(self.execute_word)
        self.btn_xl.clicked.connect(self.execute_xlsx)

    def execute_word(self):
        document_path = Path(__file__).parent / "template.docx"
        doc = DocxTemplate(document_path)
        context = {"Corp": self.lineEdit.text(),
                   "First_last_name": self.lineEdit_2.text(),
                   "Title": self.lineEdit_3.text(),
                   "Own_email": self.lineEdit_4.text(),
                   "Phone2": self.lineEdit_5.text(),
                   "Phone1": self.lineEdit_9.text(),
                   "Phone3": self.lineEdit_10.text(),
                   "Street": self.lineEdit_6.text(),
                   "City": self.lineEdit_7.text(),
                   "Website": self.lineEdit_8.text()}
        doc.render(context)
        doc.save(Path(__file__).parent / "result.docx")
        os.system('start result.docx')

    def execute_xlsx(self):
        fn = 'template.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['A1'] = self.lineEdit.text() + "    "
        ws['A3'] = "   " + self.lineEdit_2.text()
        ws['A4'] = "   " + self.lineEdit_3.text()
        ws['A6'] = "   " + self.lineEdit_4.text()
        ws['E8'] = self.lineEdit_9.text()
        ws['E9'] = self.lineEdit_5.text()
        ws['E10'] = self.lineEdit_10.text()
        ws['E5'] = self.lineEdit_6.text()
        ws['E6'] = self.lineEdit_7.text()
        ws['A9'] = "   " + self.lineEdit_8.text()

        wb.save(Path(__file__).parent / "result.xlsx")
        wb.close()
        os.system('start result.xlsx')


def main():
    app = QApplication(sys.argv)
    window = Main()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
